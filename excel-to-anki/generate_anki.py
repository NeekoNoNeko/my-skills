"""
四级真题核心词 → Anki .apkg 生成器
- 批量并行调用 edge-tts 为每个单词生成美式发音 MP3
- 仅处理 known 列为 "Selected" 的行
- 排除 known 列和随机排序列
- 生成带音频的 .apkg 文件
"""

import asyncio
import hashlib
import os
import re
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path

import edge_tts
import genanki
import openpyxl

# ─── 配置 ───
EXCEL_PATH = "四级真题核心词.xlsx"   # 输入
AUDIO_DIR = "audio"                  # 音频缓存目录
OUTPUT_PATH = "四级真题核心词.apkg"  # 输出
VOICE = "en-US-JennyNeural"          # TTS 音色
BATCH_SIZE = 8                       # 并行数
MAX_RETRIES = 2

FIELD_NAMES = [
    "单词", "词性", "释义", "美式音标", "英式音标",
    "短语", "短语释义", "例句", "例句释义", "音频",
]

# ─── 数据模型 ───
@dataclass
class WordRow:
    单词: str = ""
    词性: str = ""
    释义: str = ""
    美式音标: str = ""
    英式音标: str = ""
    短语: str = ""
    短语释义: str = ""
    例句: str = ""
    例句释义: str = ""
    音频文件: str = ""
    音频: str = ""

def read_excel(path: str) -> list[WordRow]:
    """读取 Excel，仅保留 known=Selected 的行，排除 known/随机排序两列"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    rows: list[WordRow] = []
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, col_idx["known"]).value or "").strip() != "Selected":
            continue
        word = str(ws.cell(r, col_idx["单词"]).value or "").strip()
        if not word:
            continue
        row = WordRow()
        for fn in WordRow.__dataclass_fields__:
            if fn in ("音频文件", "音频"):
                continue
            val = str(ws.cell(r, col_idx[fn]).value or "")
            setattr(row, fn, val)

        h = hashlib.md5(word.encode()).hexdigest()[:8]
        safe = re.sub(r"[^\w]", "_", word)[:30] or "word"
        row.音频文件 = f"word_{safe}_{h}.mp3"
        row.音频 = f"[sound:{row.音频文件}]"
        rows.append(row)
    return rows

# ─── TTS 音频生成（进程池并行）─
def gen_single_audio(word: str, filepath: str) -> tuple[str, str]:
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            asyncio.run(edge_tts.Communicate(word, voice=VOICE).save(filepath))
            if os.path.getsize(filepath) > 100:
                return filepath, "ok"
        except Exception as e:
            if attempt < MAX_RETRIES:
                time.sleep(0.3)
            else:
                return filepath, f"failed: {e}"
    return filepath, "failed"

def generate_audio(rows: list[WordRow], audio_dir: str) -> None:
    os.makedirs(audio_dir, exist_ok=True)
    pending = [(r.单词, os.path.join(audio_dir, r.音频文件)) for r in rows
               if not (os.path.exists(os.path.join(audio_dir, r.音频文件))
                       and os.path.getsize(os.path.join(audio_dir, r.音频文件)) > 100)]
    if not pending:
        print(f"All audio cached ({len(rows)} words)")
        return

    total = len(pending)
    print(f"Generating {total} audio files ({BATCH_SIZE} parallel)...")
    start = time.time()
    done = failed = 0
    with ProcessPoolExecutor(max_workers=BATCH_SIZE) as executor:
        fut_map = {executor.submit(gen_single_audio, w, p): (w, p) for w, p in pending}
        for fut in as_completed(fut_map):
            _, status = fut.result()
            done += 1
            failed += (status != "ok")
            if done % 100 == 0 or done == total:
                print(f"  [{done}/{total}] ({time.time()-start:.0f}s)")
    print(f"Done: {total-failed}/{total} OK, {failed} failed, {time.time()-start:.0f}s")

# ─── Anki 卡组 ───
CARD_CSS = """
.card { font-family: "Microsoft YaHei", "Noto Sans SC", "Helvetica Neue", sans-serif; font-size: 18px; text-align: center; color: #333; background: #fff; line-height: 1.6; padding: 20px; }
.word { font-size: 36px; font-weight: bold; color: #2c3e50; margin: 20px 0 10px; }
.phonetic { font-size: 16px; color: #7f8c8d; margin: 5px 0; }
.pos { font-size: 16px; color: #e67e22; margin: 5px 0; }
.meaning { font-size: 22px; color: #2c3e50; margin: 10px 0 20px; }
.section { margin: 15px 0; }
.section-title { font-size: 14px; color: #95a5a6; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 5px; }
.section-content { font-size: 16px; color: #555; }
"""
FRONT = """<div class="card">{{音频}}<div class="word">{{单词}}</div><div class="phonetic">{{英式音标}}  {{美式音标}}</div><div class="pos">{{词性}}</div></div>"""
BACK = """<div class="card">{{音频}}<div class="word">{{单词}}</div><div class="phonetic">{{英式音标}}  {{美式音标}}</div><div class="pos">{{词性}}</div><div class="meaning">{{释义}}</div>{{#短语}}<div class="section"><div class="section-title">📖 短语</div><div class="section-content">{{短语}}</div></div>{{/短语}}{{#短语释义}}<div class="section"><div class="section-title">📝 短语释义</div><div class="section-content">{{短语释义}}</div></div>{{/短语释义}}{{#例句}}<div class="section"><div class="section-title">💬 例句</div><div class="section-content">{{例句}}</div></div>{{/例句}}{{#例句释义}}<div class="section"><div class="section-title">🔍 例句释义</div><div class="section-content">{{例句释义}}</div></div>{{/例句释义}}</div>"""

def build_apkg(rows: list[WordRow], audio_dir: str, output: str) -> str:
    model_id = int.from_bytes(hashlib.md5(b"model").digest()[:4], "big") % (2**31 - 1)
    model = genanki.Model(model_id, "四级真题核心词",
        fields=[{"name": f} for f in FIELD_NAMES],
        templates=[{"name": "Card", "qfmt": FRONT, "afmt": BACK}], css=CARD_CSS)
    deck = genanki.Deck(
        int.from_bytes(hashlib.md5(b"deck").digest()[:4], "big") % (2**31 - 1),
        "四级真题核心词")
    for r in rows:
        deck.add_note(genanki.Note(model=model, fields=[getattr(r, f) for f in FIELD_NAMES]))
    media = [os.path.join(audio_dir, r.音频文件) for r in rows if os.path.exists(os.path.join(audio_dir, r.音频文件))]
    genanki.Package(deck, media_files=media).write_to_file(output)
    return output

if __name__ == "__main__":
    rows = read_excel(EXCEL_PATH)
    print(f"Read {len(rows)} words")
    generate_audio(rows, AUDIO_DIR)
    path = build_apkg(rows, AUDIO_DIR, OUTPUT_PATH)
    print(f"Created {path} ({os.path.getsize(path)/1024/1024:.1f} MB, {len(rows)} cards)")
